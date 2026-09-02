import openpyxl
from app.parsers.base import BaseParser
from app.schemas.source_item import SourceItem
from app.schemas.extracted_document import ExtractedDocument

class XLSXParser(BaseParser):
    def supports(self, content_type: str) -> bool:
        return content_type == "xlsx"

    def parse(self, source_item: SourceItem) -> ExtractedDocument:
        wb = openpyxl.load_workbook(source_item.raw_path, data_only=True)
        
        full_text = []
        sheets_data = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows_data = []
            
            for row in sheet.iter_rows(values_only=True):
                # Filter out None values and convert to string
                row_values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if row_values:
                    row_str = " | ".join(row_values)
                    rows_data.append(row_str)
                    full_text.append(row_str)
            
            sheets_data.append({
                "name": sheet_name,
                "rows": rows_data
            })
            
        parser_metadata = {
            "sheet_names": wb.sheetnames,
            "sheets": sheets_data
        }
        
        return ExtractedDocument(
            source_item=source_item,
            raw_text="\n".join(full_text),
            parser_metadata=parser_metadata
        )
